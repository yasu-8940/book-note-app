import streamlit as st
import os
import json
import pickle
import base64
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from datetime import datetime
from google.auth.transport.requests import Request
import requests
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io, requests
# from __future__ import print_function
from pathlib import Path
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload, MediaIoBaseDownload

# 🔹 Google Drive API のスコープ
SCOPES = ['https://www.googleapis.com/auth/drive.file']

# =========================================================
# Google Drive サービス取得
# =========================================================

def get_gdrive_service():

    """Google Drive API サービスを返す（OAuth 方式）"""
    creds = None

    # Render 環境: TOKEN_PICKLE_B64 を優先
    if "TOKEN_PICKLE_B64" in os.environ:
        token_bytes = base64.b64decode(os.environ["TOKEN_PICKLE_B64"])
        creds = pickle.loads(token_bytes)

    # ローカル環境: token.pickle を利用
    elif os.path.exists("token.pickle"):
        with open("token.pickle", "rb") as token:
            creds = pickle.load(token)

    # 期限切れならリフレッシュ
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())

    if not creds:
        raise FileNotFoundError("token.pickle が見つかりません。")

    return build("drive", "v3", credentials=creds)

# =========================================================
# Excel ファイル作成（表紙画像付き）
# =========================================================
def create_excel_with_image(book, comment, base_xlsx_bytes=None, filename="book_note.xlsx"):

    # 既存ファイルのBytesが来ていればそれをベースに、無ければ新規
    if base_xlsx_bytes:
        wb = load_workbook(filename=BytesIO(base_xlsx_bytes))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['登録日','書名','著者','出版社','出版日','概要','感想','表紙'])
    
    # 実際に値が入っている最後の行番号を探す
    last_row = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row):
            last_row = idx
            print(last_row)

    # 追加行はその次の行
    next_row = last_row + 1

    print("ws.max_row (openpyxl認識):", ws.max_row)
    print("last_row (実データ判定):", last_row)
    print("row_num (これから書き込む):", next_row)

    today = datetime.today().strftime("%Y-%m-%d")
    ws.append([
        today,
        book.get('title',''),
        book.get('authors',''),
        book.get('publisher',''),
        book.get('publishedDate',''),
        book.get('description',''),
        comment,
        ''
    ])

    # 表紙（PIL→openpyxl 直接渡し）
    if book.get('thumbnail'):
        r = requests.get(book['thumbnail'], timeout=10)
        r.raise_for_status()
        img_pil = Image.open(BytesIO(r.content))
        excel_img = XLImage(img_pil)
        ws.add_image(excel_img, f"H{ws.max_row}")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# =========================================================
# GoogleAPPから本を探す
# =========================================================

def search_books_google_books(title):
    url = 'https://www.googleapis.com/books/v1/volumes'
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }
    params = {
        'q': title,
        'maxResults': 10,
        'printType': 'books',
        'langRestrict': 'ja',
        # 'key': 'AIzaSyA0gXAcX6_aShRD4eKPA6ag_4QBTQtvC0w'
    }
    try:
        response = requests.get(url, headers=headers, params=params)


        data = response.json()

        if 'items' not in data:
            st.warning("⚠️ 書籍が見つかりませんでした。")
            return []

        results = []

        for item in data['items']:
            info = item['volumeInfo']
            results.append({
                'title': info.get('title', ''),
                'authors': ', '.join(info.get('authors', [])),
                'publishedDate': info.get('publishedDate', ''),
                'description': info.get('description', ''),
                'thumbnail': info.get('imageLinks', {}).get('thumbnail', ''),
                'publisher': info.get('publisher', ''),
         })
        return results

    except Exception as e:
        st.error(f"エラーが発生しました: {e}")
        return []
    
# =========================================================
# Google Drive 上書き保存関数
# =========================================================

def upload_to_drive(excel_data, folder_id, filename="book_note.xlsx"):
    service = get_gdrive_service()

    # 既存ファイルがあるか検索
    query = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get("files", [])

    media = MediaIoBaseUpload(
        io.BytesIO(excel_data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False
    )

    if items:
        # 更新（上書き）
        file_id = items[0]["id"]
        updated_file = service.files().update(
            fileId=file_id,
            media_body=media,
            fields="id, name, modifiedTime, version"
        ).execute()
        return updated_file["id"], updated_file["modifiedTime"], updated_file.get("version")
    else:
        # 新規作成
        file_metadata = {"name": filename, "parents": [folder_id]}
        new_file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, name, modifiedTime, version"
        ).execute()
        return new_file["id"], new_file["modifiedTime"], new_file.get("version")

# =========================================================
# Driveからファイルをダウンロード
# =========================================================

def download_from_drive(folder_id, filename="book_note.xlsx"):
    service = get_gdrive_service()

    # Drive上にファイルがあるか検索
    query = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get("files", [])

    if not items:
        return None  # ファイルがまだ存在しない

    file_id = items[0]["id"]
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.seek(0)    
    return fh.getvalue()

# =========================================================
# Streamlit アプリ
# =========================================================
st.title("📚 読書ノート:シリーズ対応版（Google Books API）")

search_query = st.text_input("書名を入力してください（シリーズ名もOK）：")

if st.button("候補を検索"):
    try:
        results = search_books_google_books(search_query)
        st.session_state['search_results'] = results
    except Exception as e:
        st.error(f"⚠️ エラーが発生しました: {e}")
    # results = search_books_google_books(search_query)
    # st.session_state['search_results'] = results

# 結果がある場合のみ処理
if 'search_results' in st.session_state and st.session_state['search_results']:
    results = st.session_state['search_results']
    options = [f"{book['title']} / {book['authors']}" for book in results]
    
    # 🔑 radioボタンは毎回再描画されるように
    selected = st.radio("候補から選んでください：", options, key=f"book_radio_{search_query}")
    selected_book = results[options.index(selected)]

    # 詳細表示
    st.subheader(selected_book['title'])
    st.write(f"著者: {selected_book['authors']}")
    st.write(f"出版社: {selected_book['publisher']}")
    st.write(f"出版日: {selected_book['publishedDate']}")
    st.write("概要:")
    st.write(selected_book['description'])

    if selected_book['thumbnail']:
        st.image(selected_book['thumbnail'], caption='表紙画像', width=150)

    # 感想入力
    st.markdown("---")
    comment = st.text_area("📖 感想を入力してください:")

    # ✅ Streamlit 側のダウンロードボタン（呼び出し例）
    if st.button("Excelでダウンロード（表紙付き）"):
        excel_data = create_excel_with_image(selected_book, comment)
        st.download_button(
            label="📥 Excelファイルをダウンロード",
            data=excel_data,
            file_name="book_note.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ✅ Streamlit Google Drive保存ボタン 
    if st.button("📤 Google Driveに保存（上書き）"):

        folder_id = "1CP9mzd7dOaPG9Fj88vY6OYSKwl7el1XT"

        # 1. Driveから既存Excelをダウンロード
        existing_bytes = download_from_drive(folder_id, "book_note.xlsx")

        # 既存状況のデバッグ表示
        if existing_bytes:
            wb_tmp = load_workbook(filename=BytesIO(existing_bytes))  # ← file= を使う
            st.info(f"DEBUG: 今の最終行（保存前）: {wb_tmp.active.max_row}")
        else:
            st.info("DEBUG: 既存ファイルなし（新規作成）")

        # 2. 行を追加
        st.write("DEBUG 選択書籍:", selected_book.get('title'))
        st.write("DEBUG 感想:", comment)
        excel_data = create_excel_with_image(selected_book, comment, base_xlsx_bytes=existing_bytes)

        # 3. Drive へ保存（結果も確認表示）
        file_id, modified, version = upload_to_drive(excel_data, folder_id, filename="book_note.xlsx")
        st.success(f"✅ Google Driveに保存しました！\nID: {file_id}\n更新時刻: {modified}\n版: {version}")
        st.caption(f"https://drive.google.com/file/d/{file_id}/view")
